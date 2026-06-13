"""Post-Processor: randomisiert die Uhrzeiten in den simulierten Excel-Dateien
unter `simulation_output/`, damit nicht alle Messungen eines Tages dieselbe
Uhrzeit haben.

- Hebt den Blattschutz auf (Passwort: hexhex)
- Liest Spalte „Datum" aus dem Spaltenkopf in Zeile 9
- Generiert pro Datumsgruppe sequenzielle Zeitstempel innerhalb einer
  Arbeitsschicht-Fensters (06:00–18:00) mit Jitter, sodass die Reihenfolge
  der Zeilen erhalten bleibt
- Re-aktiviert den Blattschutz und speichert

Aufruf:
    python scripts/randomize_times.py
"""
from __future__ import annotations

import random
import sys
from pathlib import Path
from collections import defaultdict

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl

from src.excel.creator import SHEET_PROTECTION_PASSWORD

OUTPUT_DIR = ROOT / "simulation_output"
HEADER_ROW = 9
DATUM_COL_NAME = "Datum"

# Arbeitsfenster (in Sekunden seit Mitternacht)
WORK_START_SEC = 6 * 3600     # 06:00
WORK_END_SEC = 18 * 3600      # 18:00
SPAN_SEC = WORK_END_SEC - WORK_START_SEC

# Reproduzierbar: gleicher Seed → gleiche Zufallsverteilung
random.seed(42)


def randomize_file(path: Path) -> tuple[int, int]:
    """Randomisiert die Uhrzeiten in einer Datei. Gibt (n_rows, n_dates) zurück."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Blattschutz aufheben
    ws.protection.sheet = False

    # Datum-Spalte aus Zeile 9 finden
    datum_col = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(HEADER_ROW, c).value == DATUM_COL_NAME:
            datum_col = c
            break
    if datum_col is None:
        wb.close()
        return (0, 0)

    # Zeilen pro Datum sammeln (Reihenfolge bewahren)
    rows_by_date: dict[str, list[int]] = defaultdict(list)
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        v = ws.cell(r, datum_col).value
        if v is None:
            continue
        s = str(v).strip()
        if len(s) < 10:
            continue
        date_part = s[:10]  # "YYYY-MM-DD"
        rows_by_date[date_part].append(r)

    total_rows = sum(len(rs) for rs in rows_by_date.values())

    for date_part, row_indices in rows_by_date.items():
        n = len(row_indices)
        # Schrittweite: gleichmäßig über das Arbeitsfenster, mit etwas Spielraum
        step = SPAN_SEC / (n + 1)
        # Etwas zufälliger Versatz am Anfang
        current = WORK_START_SEC + random.uniform(0, step)
        # Pro Zeile: aktuellen Zeitstempel + Jitter (±25 % von step), monoton steigend
        for row_idx in row_indices:
            jitter = random.uniform(-step * 0.25, step * 0.25)
            t = current + jitter
            t = max(WORK_START_SEC, min(WORK_END_SEC - 1, t))
            current += step

            h = int(t // 3600)
            m = int((t % 3600) // 60)
            s = int(t % 60)
            ws.cell(row_idx, datum_col, value=f"{date_part} {h:02d}:{m:02d}:{s:02d}")

    # Blattschutz wieder einschalten
    ws.protection.sheet = True
    ws.protection.password = SHEET_PROTECTION_PASSWORD

    wb.save(path)
    wb.close()
    return (total_rows, len(rows_by_date))


def main() -> None:
    files = sorted(OUTPUT_DIR.glob("*.xlsx"))
    if not files:
        print(f"Keine Dateien in {OUTPUT_DIR}")
        return
    for f in files:
        n_rows, n_dates = randomize_file(f)
        print(f"  {f.name}: {n_rows} Zeilen über {n_dates} Tag(e) randomisiert")


if __name__ == "__main__":
    main()
