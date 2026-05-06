"""Simulation: schreibt für Produkt _REF31842.3 fünf Excel-Dateien
(eine pro IPC), genauso wie sie die App über die UI erzeugen würde.

Quellen:
- IPC1-IPC4: handschriftliche PDF-Chargenkontrollblätter (transkribiert weiter unten)
- IPC5:     `F:\\Vorlage_3.960_St - LOT 265391.xlsm` (programmatisch gelesen)

Aufruf:
    python scripts/simulate_inputs.py

Ergebnis: 5 .xlsx-Dateien in `simulation_output/`.
"""
from __future__ import annotations

import shutil
import sys
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl

from src.config.process_config import (
    ProcessConfig,
    get_info_header_fields,
    load_product_config,
)
from src.excel.creator import create_measurement_file, write_info_header
from src.excel.writer import write_measurement_rows


# --- Konstanten ---
PRODUCT_JSON = ROOT / "data" / "products" / "REF31842_3.json"
OUTPUT_DIR = ROOT / "simulation_output"

LOT = "265391"
FA_NR = "QFA2600240"
VERWENDBAR = "31.03.2031"
SHIFT = "1"

PRODUCT_ID = "_REF31842.3"
PRODUCT_DISPLAY = (
    "_REF 31842.3 - Excelsior Sponge Strip 8,5 x 5,1 x 0,8 mm, "
    "non-sterile, bag 5000 pcs."
)

MESSMITTEL_PER_IPC = {
    "IPC1_Probenfertigung": "",
    "IPC2_Vorschneiden": "Arbeitsmessgerät 50-2",
    "IPC3_Walzen": "230106-10",
    "IPC4_Fertigschneiden": "11209250",
    "IPC5_Stanzen": "Schieblehre 2153, Schichtdicke 11008328",
}

EXCEL_IPC5 = Path("F:/Vorlage_3.960_St - LOT 265391.xlsm")


# --- Quelldaten aus den PDFs (handtranskribierte repräsentative Auszüge) ---

# IPC1: (rollencharge, rollen_nr, muster_gefertigt, datum_str, signer)
IPC1_DATA = [
    ("252619", "268C", "Ja", "28.04.2026", "Walter"),
    ("252619", "270C", "Ja", "28.04.2026", "Walter"),
    ("252668", "301A", "Ja", "04.05.2026", "Walter"),
    ("252668", "302A", "Ja", "04.05.2026", "Walter"),
    ("252668", "300A", "Ja", "04.05.2026", "Walter"),
    ("252668", "309A", "Ja", "04.05.2026", "Walter"),
    ("252667", "241C", "Ja", "04.05.2026", "Walter"),
    ("252668", "304A", "Ja", "04.05.2026", "Walter"),
    ("252668", "305A", "Ja", "04.05.2026", "Walter"),
    ("252668", "306A", "Ja", "04.05.2026", "Walter"),
    ("252668", "363A", "Ja", "04.05.2026", "Walter"),
    ("252667", "242C", "Ja", "04.05.2026", "Walter"),
    ("252667", "243C", "Ja", "04.05.2026", "Walter"),
    ("252667", "235C", "Ja", "04.05.2026", "Walter"),
    ("252667", "237C", "Ja", "04.05.2026", "Walter"),
    ("252667", "240C", "Ja", "04.05.2026", "Walter"),
    ("252667", "236C", "Ja", "04.05.2026", "Walter"),
    ("252667", "238C", "Ja", "04.05.2026", "Walter"),
]

# IPC2: (rollencharge, rolle_nr, ask_value(int), [b1..b5], datum_str, signer)
IPC2_DATA = [
    ("175268", "301A", 1456, [72, 72, 72, 72, 72], "05.05.2026", "Walter"),
    ("175268", "301A", 1530, [72, 72, 72, 72, 72], "05.05.2026", "Walter"),
    ("175268", "305A", 1827, [72, 72, 72, 72, 72], "05.05.2026", "Walter"),
    ("175268", "303A", 1516, [72, 72, 72, 72, 72], "05.05.2026", "Walter"),
    ("175268", "300A", 1481, [72, 72, 72, 72, 72], "05.05.2026", "Walter"),
    ("1252665", "282A", 1457, [71, 71, 71, 71, 71], "03.04.2026", "Zelberg"),
    ("1252665", "281A", 1470, [71, 71, 71, 71, 71], "03.04.2026", "Zelberg"),
    ("1252665", "283A", 1552, [71, 71, 71, 71, 71], "03.04.2026", "Zelberg"),
    ("1252665", "280A", 1565, [71, 71, 71, 71, 71], "03.04.2026", "Zelberg"),
    ("1252665", "277A", 1531, [71, 71, 71, 71, 71], "13.04.2026", "Walter"),
    ("1252665", "273A", 1501, [71, 71, 71, 71, 71], "13.04.2026", "Walter"),
    ("1252665", "276A", 1486, [71, 71, 71, 71, 71], "13.04.2026", "Walter"),
    ("1252665", "274A", 1530, [71, 71, 71, 71, 71], "13.04.2026", "Walter"),
    ("1242630", "177B", 1524, [71, 71, 71, 71, 71], "13.04.2026", "Walter"),
    ("1252666", "245A", 1501, [72, 71, 71, 71, 72], "13.04.2026", "Lol"),
    ("1252666", "245A", 1555, [72, 71, 71, 71, 72], "13.04.2026", "Lol"),
    ("1252666", "248A", 1567, [72, 71, 72, 71, 72], "13.04.2026", "Lol"),
    ("1252666", "247A", 1585, [72, 71, 72, 71, 72], "13.04.2026", "Lol"),
    ("1252666", "248A", 1500, [72, 71, 72, 71, 72], "13.04.2026", "Lol"),
]

# IPC3: (rollen_nr, [(L1,R1)..(L5,R5)], datum_str, signer)
IPC3_DATA = [
    ("237C", [(0.72, 0.73), (0.73, 0.74), (0.72, 0.74), (0.72, 0.74), (0.72, 0.72)],
     "05.05.2026", "Lol"),
    ("304A", [(0.72, 0.70), (0.72, 0.72), (0.72, 0.72), (0.73, 0.72), (0.73, 0.70)],
     "05.05.2026", "Lol"),
    ("305A", [(0.73, 0.74), (0.74, 0.74), (0.72, 0.72), (0.73, 0.73), (0.72, 0.72)],
     "05.05.2026", "Lol"),
    ("300A", [(0.75, 0.73), (0.73, 0.74), (0.73, 0.74), (0.73, 0.72), (0.71, 0.72)],
     "06.05.2026", "Walter"),
    ("304A", [(0.72, 0.75), (0.73, 0.74), (0.73, 0.74), (0.72, 0.72), (0.71, 0.72)],
     "06.05.2026", "Walter"),
    ("306A", [(0.72, 0.72), (0.73, 0.75), (0.72, 0.72), (0.72, 0.72), (0.72, 0.71)],
     "06.05.2026", "Walter"),
    ("303A", [(0.72, 0.71), (0.72, 0.72), (0.72, 0.72), (0.72, 0.73), (0.72, 0.71)],
     "06.05.2026", "Walter"),
    ("274C", [(0.70, 0.72), (0.70, 0.72), (0.70, 0.71), (0.71, 0.73), (0.71, 0.72)],
     "29.04.2026", "Walter"),
    ("162A", [(0.70, 0.72), (0.71, 0.71), (0.72, 0.72), (0.73, 0.72), (0.73, 0.73)],
     "30.04.2026", "Naumann"),
    ("152A", [(0.70, 0.73), (0.73, 0.73), (0.72, 0.73), (0.73, 0.73), (0.72, 0.73)],
     "30.04.2026", "Naumann"),
    ("163A", [(0.72, 0.71), (0.72, 0.72), (0.73, 0.72), (0.72, 0.72), (0.72, 0.72)],
     "30.04.2026", "Naumann"),
    ("270C", [(0.72, 0.70), (0.73, 0.71), (0.72, 0.73), (0.72, 0.73), (0.71, 0.72)],
     "30.04.2026", "Naumann"),
    ("218C", [(0.73, 0.70), (0.72, 0.70), (0.71, 0.72), (0.71, 0.71), (0.71, 0.72)],
     "30.04.2026", "Naumann"),
    ("264A", [(0.73, 0.72), (0.71, 0.71), (0.71, 0.71), (0.71, 0.72), (0.72, 0.72)],
     "05.05.2026", "Lol"),
]

# IPC4: (rollen_nr, bahn, [n1..n7 Breite mm], datum_str, signer)
IPC4_DATA = [
    ("304A", "3", [8.5, 8.4, 8.4, 8.5, 8.5, 8.5, 8.5], "06.05.2026", "Naumann"),
    ("304A", "2", [8.5, 8.6, 8.6, 8.6, 8.5, 8.5, 8.5], "06.05.2026", "Naumann"),
    ("304A", "1", [8.5, 8.6, 8.6, 8.6, 8.5, 8.5, 8.5], "06.05.2026", "Naumann"),
    ("304A", "5", [8.5, 8.6, 8.6, 8.6, 8.5, 8.5, 8.5], "06.05.2026", "Naumann"),
    ("300A", "4", [8.5, 8.5, 8.5, 8.4, 8.5, 8.5, 8.5], "06.05.2026", "Naumann"),
    ("300A", "3", [8.5, 8.5, 8.5, 8.4, 8.5, 8.5, 8.5], "06.05.2026", "Naumann"),
    ("300A", "2", [8.5, 8.5, 8.5, 8.4, 8.5, 8.5, 8.5], "06.05.2026", "Naumann"),
    ("300A", "1", [8.5, 8.5, 8.5, 8.4, 8.5, 8.5, 8.5], "06.05.2026", "Naumann"),
    ("305A", "3", [8.5, 8.5, 8.5, 8.4, 8.5, 8.5, 8.5], "06.05.2026", "Naumann"),
    ("301A", "3", [8.4, 8.5, 8.5, 8.5, 8.5, 8.4, 8.5], "06.05.2026", "Naumann"),
]


# --- Helpers ---

def parse_de_date(s: str) -> date:
    parts = s.strip().split(".")
    return date(int(parts[2]), int(parts[1]), int(parts[0]))


def make_datum(datum_str: str) -> str:
    """App schreibt Datum als 'YYYY-MM-DD HH:MM:SS'."""
    if not datum_str:
        return ""
    try:
        d = parse_de_date(datum_str)
    except (ValueError, IndexError):
        return datum_str
    return d.strftime("%Y-%m-%d 12:00:00")


def setup_file(process: ProcessConfig, dt: date, messmittel: str) -> Path:
    filepath = create_measurement_file(
        process, PRODUCT_ID, OUTPUT_DIR, LOT, FA_NR, SHIFT, dt,
    )
    extra_info: list[tuple[str, str]] = []
    for fd in get_info_header_fields(process):
        if fd.display_name == "FA-Nr.":
            value = FA_NR
        elif fd.display_name == "LOT Nr.":
            value = LOT
        elif fd.display_name == "Verwendbarkeitsdatum":
            value = VERWENDBAR
        elif fd.display_name == "Messmittel":
            value = messmittel
        else:
            value = ""
        extra_info.append((f"{fd.display_name}:", value))
    write_info_header(
        filepath=filepath,
        product_name=PRODUCT_DISPLAY,
        process_name=process.display_name,
        shift=SHIFT,
        dt=dt,
        extra_info=extra_info,
    )
    return filepath


# --- Pro IPC ---

def simulate_ipc1(product) -> Path:
    process = next(p for p in product.processes if p.template_id == "IPC1_Probenfertigung")
    dt = min(parse_de_date(r[3]) for r in IPC1_DATA)
    filepath = setup_file(process, dt, MESSMITTEL_PER_IPC[process.template_id])
    rows = []
    for rollencharge, rollen_nr, muster, datum_str, signer in IPC1_DATA:
        rows.append({
            "Rollencharge": rollencharge,
            "Rollen Nr.": rollen_nr,
            "Muster gefertigt": muster,
            "Bemerkungen": "n/a",
            "Datum": make_datum(datum_str),
            "Unterschrift": signer,
        })
    write_measurement_rows(filepath, rows)
    return filepath


def simulate_ipc2(product) -> Path:
    process = next(p for p in product.processes if p.template_id == "IPC2_Vorschneiden")
    dt = min(parse_de_date(r[4]) for r in IPC2_DATA)
    filepath = setup_file(process, dt, MESSMITTEL_PER_IPC[process.template_id])
    rows = []
    for rollencharge, rolle_nr, ask, breites, datum_str, signer in IPC2_DATA:
        ask_choice = "Ja" if ask >= 1200 else "Nein"
        for nutzen, b in enumerate(breites, 1):
            rows.append({
                "Rollencharge": rollencharge,
                "Rolle Nr.": rolle_nr,
                "Bahn": nutzen,
                "ASK Soll >=1200%": ask_choice,
                "Breite": b,
                "Bemerkungen": "n/a",
                "Datum": make_datum(datum_str),
                "Unterschrift": signer,
            })
    write_measurement_rows(filepath, rows)
    return filepath


def simulate_ipc3(product) -> Path:
    process = next(p for p in product.processes if p.template_id == "IPC3_Walzen")
    dt = min(parse_de_date(r[2]) for r in IPC3_DATA)
    filepath = setup_file(process, dt, MESSMITTEL_PER_IPC[process.template_id])
    rows = []
    for rolle_nr, bahn_data, datum_str, signer in IPC3_DATA:
        for nutzen, (links, rechts) in enumerate(bahn_data, 1):
            rows.append({
                "Rolle Nr.": rolle_nr,
                "Bahn": nutzen,
                "Schichtdicke links": links,
                "Schichtdicke rechts": rechts,
                "Bemerkungen": "n/a",
                "Datum": make_datum(datum_str),
                "Unterschrift": signer,
            })
    write_measurement_rows(filepath, rows)
    return filepath


def simulate_ipc4(product) -> Path:
    process = next(p for p in product.processes if p.template_id == "IPC4_Fertigschneiden")
    dt = min(parse_de_date(r[3]) for r in IPC4_DATA)
    filepath = setup_file(process, dt, MESSMITTEL_PER_IPC[process.template_id])
    rows = []
    for rollen_nr, bahn, breites, datum_str, signer in IPC4_DATA:
        for nutzen, b in enumerate(breites, 1):
            rows.append({
                "Rollen Nr.": rollen_nr,
                "Bahn Nr.": bahn,
                "Nutzen": nutzen,
                "Breite": b,
                "Bemerkungen": "n/a",
                "Datum": make_datum(datum_str),
                "Unterschrift": signer,
            })
    write_measurement_rows(filepath, rows)
    return filepath


def simulate_ipc5(product) -> Path:
    process = next(p for p in product.processes if p.template_id == "IPC5_Stanzen")

    wb = openpyxl.load_workbook(EXCEL_IPC5, data_only=True, keep_vba=False)
    ws = wb["IPC Stanzen"]

    earliest_dt: date | None = None
    rolle_current = ""
    raw_rows = []
    for r in range(6, ws.max_row + 1):
        bag = ws.cell(r, 1).value
        roll = ws.cell(r, 2).value
        datum_unter = ws.cell(r, 3).value
        length = ws.cell(r, 5).value
        width = ws.cell(r, 6).value
        thickness = ws.cell(r, 7).value
        maschine = ws.cell(r, 8).value
        if bag is None or length is None:
            continue
        try:
            bag_int = int(bag)
        except (TypeError, ValueError):
            # Statistik-/Trenner-Zeilen ("Mittelwert", "Stdabw" etc.) überspringen
            continue
        if roll:
            rolle_current = str(roll).strip()
        datum_str = ""
        signer = ""
        if datum_unter:
            text = str(datum_unter).strip()
            sp = text.split(" ", 1)
            datum_str = sp[0]
            signer = sp[1] if len(sp) > 1 else ""
        if datum_str:
            try:
                d = parse_de_date(datum_str)
                if earliest_dt is None or d < earliest_dt:
                    earliest_dt = d
            except (ValueError, IndexError):
                pass
        raw_rows.append({
            "bag": bag_int,
            "roll": rolle_current,
            "datum": datum_str,
            "signer": signer,
            "length": float(length),
            "width": float(width),
            "thickness": float(thickness),
            "maschine": str(int(maschine)) if maschine is not None else "",
        })
    wb.close()

    if earliest_dt is None:
        earliest_dt = date(2026, 4, 9)

    filepath = setup_file(process, earliest_dt, MESSMITTEL_PER_IPC[process.template_id])

    last_datum = ""
    last_signer = ""
    rows = []
    for r in raw_rows:
        if r["datum"]:
            last_datum = r["datum"]
            last_signer = r["signer"]
        bag_no = r["bag"]
        karton_val = ((bag_no - 1) // 20) + 1  # App-Logik: 20 Beutel pro Karton
        rows.append({
            "Maschine": r["maschine"],
            "Rollen Nr.": r["roll"],
            "Bag Nr.": bag_no,
            "Karton": karton_val,
            "Länge": r["length"],
            "Breite": r["width"],
            "Schichtdicke": r["thickness"],
            "Bemerkungen": "n/a",
            "Datum": make_datum(last_datum),
            "Unterschrift": last_signer,
        })
    write_measurement_rows(filepath, rows)
    return filepath


# --- main ---

def main() -> None:
    if OUTPUT_DIR.exists():
        shutil.rmtree(OUTPUT_DIR)
    OUTPUT_DIR.mkdir(parents=True)

    product = load_product_config(PRODUCT_JSON)
    print(f"Output: {OUTPUT_DIR}")
    for fn, label in [
        (simulate_ipc1, "IPC1"),
        (simulate_ipc2, "IPC2"),
        (simulate_ipc3, "IPC3"),
        (simulate_ipc4, "IPC4"),
        (simulate_ipc5, "IPC5"),
    ]:
        fp = fn(product)
        print(f"  {label}: {fp.name}")


if __name__ == "__main__":
    main()
