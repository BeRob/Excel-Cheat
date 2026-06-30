"""Produkt- und Prozesskonfiguration aus JSON-Dateien."""

from __future__ import annotations

import json
from dataclasses import dataclass, field
from pathlib import Path


@dataclass
class FieldDef:
    id: str
    display_name: str
    type: str  # "text" | "number" | "choice" | "date"
    role: str  # "context" | "identifier" | "measurement" | "auto"
    persistent: bool = False
    spec_target: float | None = None
    spec_min: float | None = None
    spec_max: float | None = None
    options: list[str] | None = None
    optional: bool = False
    default_value: str | None = None
    # clone=true: das Feld wird je Nutzen/Bahn wiederholt und erzeugt im Wide-
    # Format je Nutzen eine eigene Excel-Spalte ("Breite Bahn 1", "Breite Bahn 2").
    # Löst das frühere group_shared ab (invertiert: group_shared=true == clone=false).
    clone: bool = False
    info_header: bool = False
    machine_scoped: bool = False


@dataclass
class ProcessConfig:
    template_id: str
    display_name: str
    fields: list[FieldDef]
    row_group_size: int | None = None
    # Quelle der Feldstruktur (Operation-Template-Name, z.B. "Schneiden").
    # None = Legacy-Prozess mit vollständiger eigener fields-Liste.
    template: str | None = None
    # Revision des Templates, aus dem fields aufgelöst wurde (für Audit/GMP).
    template_revision: int | None = None


@dataclass
class ProcessTemplate:
    """Kanonische Feldstruktur einer Operation (Superset über alle Produkte)."""

    template: str
    template_revision: int
    fields: list[FieldDef]

    def field_map(self) -> dict[str, FieldDef]:
        return {f.id: f for f in self.fields}


@dataclass
class ProductConfig:
    product_id: str
    display_name: str
    processes: list[ProcessConfig]
    output_dir: str | None = None
    revision: int = 1
    revision_history: list[dict] = field(default_factory=list)
    # --- Laufzeit-Felder (werden NICHT serialisiert, setzt load_app_config) ---
    # Vier-Augen-Freigabe-Status gegen data/products/freigaben.json
    # (Werte: src/config/freigabe.py FREIGEGEBEN/GEAENDERT/NICHT_FREIGEGEBEN).
    freigabe_status: str = "nicht freigegeben"
    # Manifest-Eintrag der Freigabe (dokument/datum/geprueft_von/...), falls vorhanden.
    freigabe: dict | None = None
    # SHA-256 der geladenen Config-Datei (für Audit-Trail und Freigabedokument).
    config_sha256: str = ""


@dataclass
class ShiftDef:
    name: str
    start_hour: int
    end_hour: int


@dataclass
class AppConfig:
    products: list[ProductConfig] = field(default_factory=list)
    shifts: list[ShiftDef] = field(default_factory=list)
    qr_prefix: str = ""
    sheet_protection_password: str = "hexhex"
    # true (Default, streng): nur freigegebene Produkte sind wählbar.
    # false (Übergangsbetrieb): nicht freigegebene Produkte erscheinen mit
    # Warnmarkierung. Schalter in app_config.json: "freigabe_pflicht".
    freigabe_pflicht: bool = True


def _parse_clone_flag(data: dict) -> bool:
    """Liest das clone-Flag, mit Abwärtskompatibilität zum alten group_shared.

    Neu: ``clone`` ist maßgeblich. Fehlt ``clone``, aber ``group_shared`` ist
    gesetzt, wird invertiert übernommen (group_shared=true bedeutete „über alle
    Nutzen geteilt", also clone=false). Ohne beide Schlüssel: clone=false."""
    if "clone" in data:
        return bool(data["clone"])
    if "group_shared" in data:
        return not bool(data["group_shared"])
    return False


def _parse_field(data: dict) -> FieldDef:
    return FieldDef(
        id=data["id"],
        display_name=data["display_name"],
        type=data.get("type", "text"),
        role=data.get("role", "measurement"),
        persistent=data.get("persistent", False),
        spec_target=data.get("spec_target"),
        spec_min=data.get("spec_min"),
        spec_max=data.get("spec_max"),
        options=data.get("options"),
        optional=data.get("optional", False),
        default_value=data.get("default_value"),
        clone=_parse_clone_flag(data),
        info_header=data.get("info_header", False),
        machine_scoped=data.get("machine_scoped", False),
    )


# Attribute, die eine dünne Config je Feld überschreiben darf — auch von
# config_writer benutzt, um beim Speichern die Overrides zurückzurechnen.
FIELD_OVERRIDE_KEYS = {
    "display_name", "type", "role", "persistent", "spec_target", "spec_min",
    "spec_max", "options", "optional", "default_value", "clone",
    "info_header", "machine_scoped",
}


def _load_json(path: Path) -> dict:
    """Liest eine JSON-Datei; im Fehlerfall wird die Datei beim Namen genannt,
    damit ein Tippfehler in einer einzelnen Config beim App-Start auffindbar ist."""
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as e:
        raise ValueError(f"Ungültiges JSON in {path}: {e}") from e
    except OSError as e:
        raise ValueError(f"{path} konnte nicht gelesen werden: {e}") from e


def parse_process_template(data: dict) -> ProcessTemplate:
    """Parst eine Operation-Template-Datei (data/process_templates/<Operation>.json)."""
    try:
        revision = int(data.get("template_revision", 1))
    except (TypeError, ValueError):
        revision = 1
    return ProcessTemplate(
        template=data["template"],
        template_revision=revision,
        fields=[_parse_field(f) for f in data.get("fields", [])],
    )


def load_process_templates(templates_dir: Path) -> dict[str, ProcessTemplate]:
    """Lädt alle Operation-Templates aus dem Verzeichnis (Schlüssel = template-Name)."""
    templates: dict[str, ProcessTemplate] = {}
    if templates_dir.exists():
        for p in sorted(templates_dir.glob("*.json")):
            if p.name.startswith("_"):
                continue
            try:
                tpl = parse_process_template(_load_json(p))
            except KeyError as e:
                raise ValueError(
                    f"Prozess-Template {p.name}: Pflichtschlüssel {e} fehlt."
                ) from e
            templates[tpl.template] = tpl
    return templates


def _apply_field_overrides(base: FieldDef, overrides: dict) -> FieldDef:
    """Erzeugt eine Kopie der FieldDef mit angewandten Override-Werten."""
    from dataclasses import replace

    changes = {k: v for k, v in overrides.items() if k in FIELD_OVERRIDE_KEYS}
    return replace(base, **changes)


def _resolve_process(
    data: dict, templates: dict[str, ProcessTemplate]
) -> ProcessConfig:
    """Löst einen dünnen Prozess (template + active_fields + Overrides) zur vollen
    ProcessConfig auf. Fällt auf Legacy-Verhalten zurück, wenn kein template gesetzt ist."""
    template_name = data.get("template")
    if not template_name or "active_fields" not in data:
        # Legacy: vollständige fields-Liste direkt in der Produkt-Config.
        return _parse_process(data)

    tpl = templates.get(template_name)
    if tpl is None:
        raise ValueError(
            f"Prozess '{data.get('template_id')}' referenziert unbekanntes "
            f"Template '{template_name}'."
        )

    tpl_fields = tpl.field_map()
    extra_fields = {
        f["id"]: _parse_field(f) for f in data.get("extra_fields", [])
    }
    field_overrides: dict[str, dict] = data.get("field_overrides", {})

    resolved: list[FieldDef] = []
    for fid in data["active_fields"]:
        if fid in tpl_fields:
            base = tpl_fields[fid]
        elif fid in extra_fields:
            base = extra_fields[fid]
        else:
            raise ValueError(
                f"Prozess '{data.get('template_id')}': aktives Feld '{fid}' weder "
                f"im Template '{template_name}' noch in extra_fields gefunden."
            )
        ov = field_overrides.get(fid)
        resolved.append(_apply_field_overrides(base, ov) if ov else base)

    return ProcessConfig(
        template_id=data["template_id"],
        display_name=data["display_name"],
        fields=resolved,
        row_group_size=data.get("row_group_size"),
        template=template_name,
        template_revision=tpl.template_revision,
    )


def _parse_process(data: dict) -> ProcessConfig:
    # template/template_revision werden auch im Legacy-Fall durchgereicht,
    # damit die Audit-Herkunft einen Editor-Roundtrip überlebt, selbst wenn
    # die Template-Datei (vorübergehend) fehlt.
    tpl_revision = data.get("template_revision")
    if tpl_revision is not None:
        try:
            tpl_revision = int(tpl_revision)
        except (TypeError, ValueError):
            tpl_revision = None
    return ProcessConfig(
        template_id=data["template_id"],
        display_name=data["display_name"],
        fields=[_parse_field(f) for f in data["fields"]],
        row_group_size=data.get("row_group_size"),
        template=data.get("template"),
        template_revision=tpl_revision,
    )


def load_product_config(
    path: Path, templates: dict[str, ProcessTemplate] | None = None
) -> ProductConfig:
    """Lädt eine Produkt-Konfiguration aus einer JSON-Datei.

    Dünne Configs (Prozesse mit ``template`` + ``active_fields``) werden gegen die
    übergebenen ``templates`` aufgelöst. Ohne ``templates`` bzw. für Legacy-Prozesse
    mit vollständiger ``fields``-Liste bleibt das Verhalten unverändert.
    """
    data = _load_json(path)
    try:
        revision = int(data.get("revision", 1))
    except (TypeError, ValueError):
        revision = 1
    history = data.get("revision_history", [])
    if not isinstance(history, list):
        history = []
    tpls = templates or {}
    try:
        return ProductConfig(
            product_id=data["product_id"],
            display_name=data["display_name"],
            processes=[_resolve_process(p, tpls) for p in data["processes"]],
            output_dir=data.get("output_dir"),
            revision=revision,
            revision_history=history,
        )
    except KeyError as e:
        raise ValueError(
            f"Produkt-Config {path.name}: Pflichtschlüssel {e} fehlt."
        ) from e
    except ValueError as e:
        raise ValueError(f"Produkt-Config {path.name}: {e}") from e


def load_app_config(
    config_path: Path,
    products_dir: Path,
    templates_dir: Path | None = None,
) -> AppConfig:
    """Lädt globale Settings, Prozess-Templates und alle Produktdateien."""
    if config_path.exists():
        global_data = _load_json(config_path)
    else:
        global_data = {}

    shifts = [
        ShiftDef(
            name=s["name"],
            start_hour=s["start_hour"],
            end_hour=s["end_hour"],
        )
        for s in global_data.get("shifts", [])
    ]

    templates = load_process_templates(templates_dir) if templates_dir else {}

    from src.config.freigabe import (
        compute_config_hash, determine_status, freigaben_path, load_freigaben,
    )

    products: list[ProductConfig] = []
    if products_dir.exists():
        freigaben = load_freigaben(products_dir)
        manifest_name = freigaben_path(products_dir).name
        for p in sorted(products_dir.glob("*.json")):
            # Manifest und Arbeitsdateien (_-Präfix) sind keine Produkt-Configs.
            if p.name == manifest_name or p.name.startswith("_"):
                continue
            product = load_product_config(p, templates)
            product.config_sha256 = compute_config_hash(p)
            product.freigabe = freigaben.get(product.product_id)
            product.freigabe_status = determine_status(
                product.freigabe, product.config_sha256, product.revision,
            )
            products.append(product)

    return AppConfig(
        products=products,
        shifts=shifts,
        qr_prefix=global_data.get("qr_prefix", ""),
        sheet_protection_password=global_data.get(
            "sheet_protection_password", "hexhex"
        ),
        freigabe_pflicht=bool(global_data.get("freigabe_pflicht", True)),
    )


def get_context_fields(process: ProcessConfig) -> list[FieldDef]:
    return [f for f in process.fields if f.role == "context"]


def get_persistent_context_fields(process: ProcessConfig) -> list[FieldDef]:
    return [f for f in process.fields if f.role == "context" and f.persistent]


def get_form_persistent_fields(process: ProcessConfig) -> list[FieldDef]:
    """Persistente Kontext-Felder, die im Formular als 'Feste Werte' gerendert werden
    (also ohne die info_header-Felder, die in der Top-Bar erscheinen)."""
    return [
        f for f in process.fields
        if f.role == "context" and f.persistent and not f.info_header
    ]


def get_info_header_fields(process: ProcessConfig) -> list[FieldDef]:
    """Felder, die in den Excel-Info-Block (Zeilen 1-5) geschrieben werden,
    nicht als Spalte in der Datenzeile."""
    return [f for f in process.fields if f.info_header]


def get_per_measurement_context_fields(process: ProcessConfig) -> list[FieldDef]:
    return [f for f in process.fields if f.role == "context" and not f.persistent]


def get_identifier_fields(process: ProcessConfig) -> list[FieldDef]:
    """Zeilen-Kennungen je Messung (Rollen-Nr., Bahn, Lfd. Nr. …) — weder reiner
    Kontext noch Messwert. Werden wie Kontext gerendert, aber als Spalte
    geschrieben und nie geklont."""
    return [f for f in process.fields if f.role == "identifier"]


def get_shared_input_fields(process: ProcessConfig) -> list[FieldDef]:
    """Felder des 'Gemeinsame Werte'-Blocks: pro-Messung-Kontext + Kennungen +
    nicht-geklonte Messwerte (einmal je Messung erfasst). info_header-Felder
    gehören in die Kopfleiste, nicht hierher."""
    return [
        f for f in process.fields
        if not f.info_header and (
            (f.role == "context" and not f.persistent)
            or f.role == "identifier"
            or (f.role == "measurement" and not f.clone)
        )
    ]


def get_measurement_fields(process: ProcessConfig) -> list[FieldDef]:
    return [f for f in process.fields if f.role == "measurement"]


def get_shared_measurement_fields(process: ProcessConfig) -> list[FieldDef]:
    """Messwerte, die einmal je Messung erfasst werden (nicht geklont)."""
    return [f for f in process.fields if f.role == "measurement" and not f.clone]


def get_clone_fields(process: ProcessConfig) -> list[FieldDef]:
    """Felder, die je Nutzen/Bahn wiederholt werden (clone=true)."""
    return [f for f in process.fields if f.clone]


def get_auto_fields(process: ProcessConfig) -> list[FieldDef]:
    return [f for f in process.fields if f.role == "auto"]


NUTZEN_FIELD_ID = "nutzen"


def is_multi_nutzen(process: ProcessConfig) -> bool:
    """True, wenn der Prozess je Messung mehrere Nutzen/Bahnen erfasst.

    Aktiv, sobald mindestens ein Feld ``clone=true`` ist. Im Wide-Format wird
    jedes Clone-Feld je Nutzen zu einer eigenen Spalte; die Anzahl Nutzen wählt
    der Bediener beim Prozessstart (Max/Default = ``row_group_size``)."""
    return any(f.clone for f in process.fields)


def get_nutzen_label(process: ProcessConfig) -> str:
    """Bezeichnung der Nutzen-Achse (z.B. 'Bahn' oder 'Nutzen'), abgeleitet aus
    dem ``nutzen``-Auto-Feld; Default 'Nutzen'."""
    for f in process.fields:
        if f.id == NUTZEN_FIELD_ID:
            return f.display_name
    return "Nutzen"


def clone_column_name(display_name: str, nutzen_label: str, index: int) -> str:
    """Spaltenname eines geklonten Feldes für Nutzen ``index`` (1-basiert),
    z.B. ('Breite', 'Bahn', 1) -> 'Breite Bahn 1'. Einzige Quelle der Wahrheit —
    Creator (Header), Writer (Validierung) und FormView (Schreiben) nutzen sie."""
    return f"{display_name} {nutzen_label} {index}"


def get_all_headers(process: ProcessConfig, nutzen_count: int = 1) -> list[str]:
    """Spaltenkopf-Namen für Zeile 9 (Info-Header-Felder ausgeschlossen).

    Im Wide-Format (Multi-Nutzen) wird jedes ``clone``-Feld zu ``nutzen_count``
    nummerierten Spalten expandiert; das ``nutzen``-Auto-Feld entfällt als Spalte
    (die Nutzen-Nr. steckt im Clone-Spaltennamen)."""
    multi = is_multi_nutzen(process)
    label = get_nutzen_label(process)
    headers: list[str] = []
    for f in process.fields:
        if f.info_header:
            continue
        if multi and f.id == NUTZEN_FIELD_ID and f.role == "auto":
            continue
        if multi and f.clone:
            for i in range(1, max(nutzen_count, 1) + 1):
                headers.append(clone_column_name(f.display_name, label, i))
        else:
            headers.append(f.display_name)
    return headers


def read_nutzen_count_from_file(filepath, process: ProcessConfig) -> int:
    """Liest beim Resume die Nutzen-Anzahl aus den vorhandenen Spaltenköpfen
    (Zeile 9), indem die nummerierten Clone-Spalten des ersten Clone-Feldes
    gezählt werden. Liefert mindestens 1. So bleibt die Anzahl je Datei stabil,
    ohne sie separat speichern zu müssen."""
    clone_fields = get_clone_fields(process)
    if not clone_fields:
        return 1
    import openpyxl

    from src.config.settings import HEADER_ROW

    label = get_nutzen_label(process)
    first = clone_fields[0].display_name
    prefix = f"{first} {label} "
    wb = openpyxl.load_workbook(filepath, read_only=True)
    try:
        ws = wb.active
        count = 0
        for col in range(1, ws.max_column + 1):
            val = ws.cell(HEADER_ROW, col).value
            if val is None:
                continue
            text = str(val)
            if text.startswith(prefix) and text[len(prefix):].strip().isdigit():
                count += 1
        return max(count, 1)
    finally:
        wb.close()


def get_field_by_id(process: ProcessConfig, field_id: str) -> FieldDef | None:
    for f in process.fields:
        if f.id == field_id:
            return f
    return None


def determine_shift(hour: int, shifts: list[ShiftDef]) -> str:
    """Bestimmt die Schicht für die aktuelle Stunde.

    Berücksichtigt Schichten über Mitternacht (z.B. 22-06).
    Fallback auf "1", wenn keine Schicht passt oder keine konfiguriert sind —
    mit Warnung im Tech-Log, weil eine falsche Schicht im Dateinamen und im
    Excel-Info-Block der Chargendokumentation landet.
    """
    for shift in shifts:
        if shift.start_hour < shift.end_hour:
            if shift.start_hour <= hour < shift.end_hour:
                return shift.name
        else:
            if hour >= shift.start_hour or hour < shift.end_hour:
                return shift.name

    import logging

    logging.getLogger("config").warning(
        "Keine Schicht für Stunde %d konfiguriert (shifts=%d) — Fallback auf '1'",
        hour, len(shifts),
    )
    return "1"
