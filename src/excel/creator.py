"""Excel-Dateien aus einer Prozesskonfiguration erstellen."""

from __future__ import annotations

import logging
import re
from datetime import datetime, date, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

from src.config.process_config import ProcessConfig, get_all_headers
from src.config.settings import HEADER_ROW
from src.excel.safe_save import save_workbook_atomic


_logger = logging.getLogger("excel.creator")


HEADER_FONT = Font(bold=True, size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrapText=True)
THIN_BORDER = Border(
    left=Side("thin"), right=Side("thin"),
    top=Side("thin"), bottom=Side("thin"),
)
INFO_LABEL_FONT = Font(bold=True, size=10)
INFO_TITLE_FONT = Font(bold=True, size=14)

SHEET_PROTECTION_PASSWORD = "hexhex"


def _apply_sheet_protection(ws, password: str = SHEET_PROTECTION_PASSWORD) -> None:
    ws.protection.sheet = True
    ws.protection.password = password


def _sanitize_for_filename(value: str) -> str:
    """Ersetzt Zeichen die in Dateinamen ungültig sind."""
    value = value.strip()
    value = re.sub(r'[\\/:*?"<>|]', "", value)
    value = re.sub(r"\s+", "_", value)
    return value or "leer"


def generate_file_name(
    lot: str,
    fa_nr: str,
    product_id: str,
    process_id: str,
    shift: str,
    dt: date,
) -> str:
    """Format: {LOT}_{FANR}_{ProductID}_{ProcessID}_{YYYY-MM-DD}_Schicht{N}.xlsx"""
    lot_s = _sanitize_for_filename(lot)
    fa_s = _sanitize_for_filename(fa_nr)
    date_str = dt.strftime("%Y-%m-%d")
    return f"{lot_s}_{fa_s}_{product_id}_{process_id}_{date_str}_Schicht{shift}.xlsx"


def get_shift_date(now: datetime, shift: str) -> date:
    """Für Schicht 3 nach Mitternacht wird der Vortag verwendet."""
    if shift == "3" and now.hour < 6:
        return (now - timedelta(days=1)).date()
    return now.date()


def find_existing_file(
    lot: str,
    fa_nr: str,
    product_id: str,
    process_id: str,
    output_dir: Path,
) -> Path | None:
    """Sucht nach einer Datei mit passendem LOT/FA-Nr/Produkt/Prozess, unabhängig von Datum/Schicht."""
    lot_s = _sanitize_for_filename(lot)
    fa_s = _sanitize_for_filename(fa_nr)
    pattern = f"{lot_s}_{fa_s}_{product_id}_{process_id}_*.xlsx"
    matches = list(output_dir.glob(pattern))
    if not matches:
        return None
    return max(matches, key=lambda p: p.stat().st_mtime)


def create_measurement_file(
    process: ProcessConfig,
    product_id: str,
    output_dir: Path,
    lot: str,
    fa_nr: str,
    shift: str,
    dt: date,
    protection_password: str = SHEET_PROTECTION_PASSWORD,
    nutzen_count: int = 1,
) -> Path:
    """Legt eine neue Excel-Datei mit Kopfzeile an und gibt den Pfad zurück.

    ``nutzen_count`` bestimmt im Wide-Format, wie viele nummerierte Spalten je
    Clone-Feld angelegt werden (einmalig je Datei festgelegt)."""
    output_dir.mkdir(parents=True, exist_ok=True)

    name = generate_file_name(lot, fa_nr, product_id, process.template_id, shift, dt)
    path = output_dir / name

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Messdaten"

    headers = get_all_headers(process, nutzen_count)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(HEADER_ROW, col_idx, header)
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

        width = max(len(header) + 4, 14)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    _apply_sheet_protection(ws, protection_password)

    save_workbook_atomic(wb, path)
    wb.close()

    return path


def count_data_rows(filepath: Path) -> int:
    """Zählt Datenzeilen (ohne Info-Block und ohne Headerzeile).

    Wirft bei Lesefehlern — der Aufrufer muss entscheiden (der Resume bricht
    ab, statt mit Zählung 0 die Prüfmuster-/Beutel-Nummerierung neu zu starten
    und doppelte Nummern in die Chargendokumentation zu schreiben)."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    try:
        ws = wb.active
        return max(ws.max_row - HEADER_ROW, 0)
    finally:
        wb.close()


def write_info_header(
    filepath: Path,
    product_name: str,
    process_name: str,
    shift: str,
    dt: date,
    extra_info: list[tuple[str, str]] | None = None,
) -> bool:
    """Schreibt den Info-Block in die Zeilen 1-8. True = erfolgreich geschrieben.

    Zeile 1 Spalte A: Produktname (fett, groß).
    Zeilen 2-4 Spalten A/B: Prozess, Schicht, Datum.
    Zeilen 2-8 Spalten C/D: zusätzliche Header-Felder (FA-Nr., LOT,
    Verwendbarkeitsdatum, Messmittel).

    Fehler werden geloggt und als False gemeldet — der Aufrufer entscheidet,
    ob der Prozessstart abgebrochen wird. Der Info-Block ist Teil des
    GMP-Records; stilles Fehlen wäre eine unvollständige Chargendokumentation."""
    try:
        wb = openpyxl.load_workbook(filepath)
    except Exception:
        _logger.exception(
            "write_info_header: %s konnte nicht geöffnet werden", filepath
        )
        return False

    try:
        ws = wb.active

        title_cell = ws.cell(1, 1, product_name)
        title_cell.font = INFO_TITLE_FONT

        core_rows = [
            ("Prozess:", process_name),
            ("Schicht:", shift),
            ("Datum:", dt.strftime("%Y-%m-%d")),
        ]

        for offset, (label, value) in enumerate(core_rows):
            row_idx = 2 + offset
            cell_label = ws.cell(row_idx, 1, label)
            cell_label.font = INFO_LABEL_FONT
            ws.cell(row_idx, 2, value)

        if extra_info:
            ws.column_dimensions["C"].width = 24
            ws.column_dimensions["D"].width = 24
            # Cap auf Zeile 12 — bis zu 11 Zusatzeinträge (FA-Nr., LOT,
            # Verwendbarkeitsdatum, Messmittel + bis zu 7 weitere Messmittel
            # bei komma-getrennter Eingabe).
            for offset, (label, value) in enumerate(extra_info):
                row_idx = 2 + offset
                if row_idx > 12:
                    break
                cell_label = ws.cell(row_idx, 3, label)
                cell_label.font = INFO_LABEL_FONT
                ws.cell(row_idx, 4, value)

        save_workbook_atomic(wb, filepath)
        return True
    except Exception:
        _logger.exception(
            "write_info_header: Schreiben fehlgeschlagen (%s)", filepath
        )
        return False
    finally:
        wb.close()
