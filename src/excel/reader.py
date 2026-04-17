"""Excel-Dateien einlesen (Header-Erkennung und vollständiger Datenabzug)."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl


@dataclass
class ExcelReadResult:
    sheet_names: list[str] = field(default_factory=list)
    headers: list[str] = field(default_factory=list)
    header_column_map: dict[str, int] = field(default_factory=dict)
    errors: list[str] = field(default_factory=list)


def read_excel_headers(
    filepath: str | Path,
    sheet_name: str | None = None,
    header_row: int = 1,
) -> ExcelReadResult:
    """Liest die Headerzeile aus einer Excel-Datei.

    Ohne `sheet_name` wird das erste Arbeitsblatt genommen. `header_row`
    ist 1-basiert.
    """
    result = ExcelReadResult()

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except FileNotFoundError:
        result.errors.append(f"Datei nicht gefunden: {filepath}")
        return result
    except PermissionError:
        result.errors.append("Datei ist gesperrt. Bitte Excel schließen.")
        return result
    except Exception as e:
        result.errors.append(f"Datei konnte nicht geöffnet werden: {e}")
        return result

    try:
        result.sheet_names = wb.sheetnames

        target_sheet = sheet_name if sheet_name else wb.sheetnames[0]
        if target_sheet not in wb.sheetnames:
            result.errors.append(f"Arbeitsblatt '{target_sheet}' nicht gefunden.")
            return result

        ws = wb[target_sheet]

        raw_headers: list = []
        for row in ws.iter_rows(
            min_row=header_row, max_row=header_row, values_only=True
        ):
            raw_headers = list(row)
            break

        if not raw_headers:
            result.errors.append("Keine Header-Zeile gefunden.")
            return result

        headers, col_map, errors = _clean_headers(raw_headers)
        result.headers = headers
        result.header_column_map = col_map
        result.errors = errors

    finally:
        wb.close()

    return result


def read_all_data(
    filepath: str | Path,
    sheet_name: str | None = None,
    header_row: int = 1,
) -> list[dict[str, Any]]:
    """Liest alle Zeilen und gibt sie als Liste von Dicts zurück."""
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception:
        return []

    try:
        target_sheet = sheet_name if sheet_name else wb.sheetnames[0]
        if target_sheet not in wb.sheetnames:
            return []

        ws = wb[target_sheet]

        raw_headers: list = []
        for row in ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True):
            raw_headers = list(row)
            break
        if not raw_headers:
            return []

        headers, _, _ = _clean_headers(raw_headers)

        data: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not any(cell is not None for cell in row):
                continue
            row_dict: dict[str, Any] = {}
            for idx, value in enumerate(row):
                if idx < len(headers):
                    row_dict[headers[idx]] = value
            data.append(row_dict)

        return data
    finally:
        wb.close()


def _clean_headers(
    raw_headers: list,
) -> tuple[list[str], dict[str, int], list[str]]:
    """Bereinigt Header: leere Zellen erzeugen Fehler, Duplikate bekommen _2, _3 ..."""
    headers: list[str] = []
    col_map: dict[str, int] = {}
    errors: list[str] = []
    seen: dict[str, int] = {}

    for idx, value in enumerate(raw_headers):
        col_num = idx + 1

        if value is None or str(value).strip() == "":
            errors.append(f"Spalte {col_num} hat keinen Header.")
            continue

        name = str(value).strip()

        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 1

        headers.append(name)
        col_map[name] = col_num

    return headers, col_map, errors
