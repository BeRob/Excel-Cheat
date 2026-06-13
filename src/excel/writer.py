"""Messzeilen an eine bestehende Excel-Datei anhängen.

GMP-Härtung:
- Atomares Speichern (Temp-Datei + Rename) — kein Korruptionsrisiko beim Absturz.
- Inter-Prozess-Lock um load→save — zwei Workstations können sich nicht
  gegenseitig Zeilen überschreiben.
- Header-Validierung gegen die Prozess-Definition — manipulierte oder
  abweichende Spaltenköpfe führen zu einem harten Fehler statt zu stillem
  Feldverlust. Werte ohne Spaltentreffer brechen das Schreiben ab; nur
  info_header-Felder (bewusst ohne Spalte) werden übersprungen.
"""

from __future__ import annotations

import logging
import os
from contextlib import contextmanager
from dataclasses import dataclass
from pathlib import Path

import openpyxl

from src.audit.file_lock import acquire_lock, release_lock
from src.config.process_config import get_all_headers, get_info_header_fields
from src.config.settings import HEADER_ROW
from src.excel.safe_save import save_workbook_atomic


_logger = logging.getLogger("excel.writer")

# Timeout für das Inter-Prozess-Lock um load→save (SMB-Mehrplatzbetrieb).
LOCK_TIMEOUT_SECONDS = 5.0


@dataclass
class WriteResult:
    success: bool = False
    row_number: int | None = None
    error: str | None = None


def _build_col_map(ws) -> dict[str, int]:
    """Liest die Spaltenüberschriften aus der Datei – verhindert Spaltenshift bei Config-Änderungen."""
    col_map = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(HEADER_ROW, col).value
        if val is not None:
            col_map[str(val)] = col
    return col_map


@contextmanager
def _excel_lock(filepath: str | Path, timeout: float = LOCK_TIMEOUT_SECONDS):
    """Serialisiert load→save über eine Lock-Datei neben der Excel-Datei.

    Liefert True (Lock gehalten) oder False (Timeout — andere Station schreibt).
    Kann die Lock-Datei selbst nicht angelegt werden (z.B. ACL), wird ohne Lock
    weitergearbeitet — ein Lock-Infrastrukturproblem darf die Messwerterfassung
    nicht lahmlegen; der Fall wird im Tech-Log vermerkt."""
    lock_path = str(filepath) + ".lock"
    try:
        fd = os.open(lock_path, os.O_RDWR | os.O_CREAT, 0o644)
    except OSError as e:
        _logger.warning("Excel-Lock-Datei nicht anlegbar (%s): %s", lock_path, e)
        yield True
        return
    try:
        if not acquire_lock(fd, timeout):
            yield False
        else:
            try:
                yield True
            finally:
                release_lock(fd)
    finally:
        try:
            os.close(fd)
        except OSError:
            pass


def _append_rows(
    filepath: str | Path,
    rows: list[dict[str, str | float | None]],
    process=None,
) -> WriteResult:
    """Gemeinsamer Kern: hängt Zeilen an, validiert Spalten, speichert atomar."""
    result = WriteResult()
    if not rows:
        result.error = "Keine Zeilen zum Schreiben."
        return result

    with _excel_lock(filepath) as locked:
        if not locked:
            result.error = (
                "Datei wird gerade von einer anderen Station beschrieben. "
                "Bitte erneut versuchen."
            )
            _logger.error("_append_rows: Lock-Timeout (%s)", filepath)
            return result

        try:
            wb = openpyxl.load_workbook(filepath)
        except PermissionError:
            result.error = "Datei ist gesperrt. Bitte Excel schließen und erneut versuchen."
            _logger.error("_append_rows: %s (%s)", result.error, filepath)
            return result
        except FileNotFoundError:
            result.error = f"Datei nicht gefunden: {filepath}"
            _logger.error("_append_rows: %s", result.error)
            return result
        except Exception as e:
            result.error = f"Datei konnte nicht geöffnet werden: {e}"
            _logger.exception("_append_rows: open failed")
            return result

        try:
            ws = wb.active
            col_map = _build_col_map(ws)

            ignore: set[str] = set()
            if process is not None:
                ignore = {f.display_name for f in get_info_header_fields(process)}
                missing = [h for h in get_all_headers(process) if h not in col_map]
                if missing:
                    result.error = (
                        "Spalten fehlen in der Datei (Header-Zeile verändert?): "
                        + ", ".join(missing)
                    )
                    _logger.error("_append_rows: %s (%s)", result.error, filepath)
                    return result

            # Werte ohne Spaltentreffer wären stiller Datenverlust → vor dem
            # Schreiben prüfen und hart abbrechen.
            skipped: set[str] = set()
            for row_data in rows:
                for display_name, value in row_data.items():
                    if value is None or display_name in ignore:
                        continue
                    if display_name not in col_map:
                        skipped.add(display_name)
            if skipped:
                result.error = (
                    "Keine Spalte für: " + ", ".join(sorted(skipped))
                    + " — Zeile wurde NICHT geschrieben."
                )
                _logger.error("_append_rows: %s (%s)", result.error, filepath)
                return result

            next_row = ws.max_row + 1
            for row_data in rows:
                for display_name, value in row_data.items():
                    col_idx = col_map.get(display_name)
                    if col_idx is not None and value is not None:
                        ws.cell(row=next_row, column=col_idx, value=value)
                next_row += 1

            save_workbook_atomic(wb, filepath)
            result.success = True
            result.row_number = next_row - 1

        except PermissionError:
            result.error = "Datei ist gesperrt. Bitte Excel schließen und erneut versuchen."
            _logger.error("_append_rows: %s (%s)", result.error, filepath)
        except Exception as e:
            result.error = f"Unerwarteter Fehler beim Schreiben: {e}"
            _logger.exception("_append_rows: write failed")
        finally:
            wb.close()

    return result


def write_measurement_row(
    filepath: str | Path,
    process,
    context_values: dict[str, str],
    measurements: dict[str, float | str | None],
    auto_values: dict[str, str | float | None],
) -> WriteResult:
    """Hängt eine Messzeile an. Spaltenposition wird aus der Datei gelesen, nicht aus process.fields."""
    all_values: dict[str, str | float | None] = {}
    all_values.update(context_values)
    all_values.update(measurements)
    all_values.update(auto_values)
    return _append_rows(filepath, [all_values], process)


def write_measurement_rows(
    filepath: str | Path,
    rows: list[dict[str, str | float | None]],
    process=None,
) -> WriteResult:
    """Hängt mehrere Messzeilen in einem Schreibvorgang an (für Multi-Nutzen)."""
    return _append_rows(filepath, rows, process)
