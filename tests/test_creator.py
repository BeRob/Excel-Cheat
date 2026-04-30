"""Tests für die Excel-Dateierstellung."""

import shutil
import tempfile
import unittest
from datetime import date, datetime
from pathlib import Path

import openpyxl

from src.config.process_config import FieldDef, ProcessConfig
from src.config.settings import HEADER_ROW
from src.excel.creator import (
    count_data_rows,
    create_measurement_file,
    find_existing_file,
    generate_file_name,
    get_shift_date,
    write_info_header,
)


def _make_process() -> ProcessConfig:
    return ProcessConfig(
        template_id="IPC1_Test",
        display_name="IPC1 Test",
        fields=[
            FieldDef(id="lot", display_name="LOT Nr.", type="text", role="context"),
            FieldDef(id="breite", display_name="Breite", type="number", role="measurement"),
            FieldDef(id="datum", display_name="Datum", type="text", role="auto"),
        ],
    )


class TestGenerateFileName(unittest.TestCase):

    def test_format(self):
        name = generate_file_name("LOT001", "FA123", "REF123", "IPC1_Test", "1", date(2026, 4, 1))
        self.assertEqual(name, "LOT001_FA123_REF123_IPC1_Test_2026-04-01_Schicht1.xlsx")

    def test_shift_3(self):
        name = generate_file_name("LOT001", "FA123", "REF123", "IPC1_Test", "3", date(2026, 3, 31))
        self.assertEqual(name, "LOT001_FA123_REF123_IPC1_Test_2026-03-31_Schicht3.xlsx")

    def test_sanitizes_spaces(self):
        name = generate_file_name("LOT 001", "FA 123", "REF123", "IPC1_Test", "1", date(2026, 4, 1))
        self.assertNotIn(" ", name)

    def test_sanitizes_slashes(self):
        name = generate_file_name("LOT/001", "FA\\123", "REF123", "IPC1_Test", "1", date(2026, 4, 1))
        self.assertNotIn("/", name)
        self.assertNotIn("\\", name)


class TestGetShiftDate(unittest.TestCase):

    def test_shift_1_normal(self):
        dt = datetime(2026, 4, 1, 10, 0)
        self.assertEqual(get_shift_date(dt, "1"), date(2026, 4, 1))

    def test_shift_3_before_midnight(self):
        dt = datetime(2026, 4, 1, 23, 0)
        self.assertEqual(get_shift_date(dt, "3"), date(2026, 4, 1))

    def test_shift_3_after_midnight(self):
        dt = datetime(2026, 4, 2, 2, 0)
        self.assertEqual(get_shift_date(dt, "3"), date(2026, 4, 1))

    def test_shift_2_not_affected(self):
        dt = datetime(2026, 4, 1, 2, 0)
        self.assertEqual(get_shift_date(dt, "2"), date(2026, 4, 1))


class TestCreateMeasurementFile(unittest.TestCase):

    def setUp(self):
        self.tmp_dir = Path(tempfile.mkdtemp())
        self.process = _make_process()

    def tearDown(self):
        shutil.rmtree(self.tmp_dir, ignore_errors=True)

    def test_creates_file(self):
        path = create_measurement_file(
            self.process, "REF123", self.tmp_dir, "LOT001", "FA123", "1", date(2026, 4, 1)
        )
        self.assertTrue(path.exists())
        self.assertEqual(path.name, "LOT001_FA123_REF123_IPC1_Test_2026-04-01_Schicht1.xlsx")

    def test_header_row(self):
        path = create_measurement_file(
            self.process, "REF123", self.tmp_dir, "LOT001", "FA123", "1", date(2026, 4, 1)
        )
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        headers = [ws.cell(HEADER_ROW, c).value for c in range(1, 4)]
        wb.close()
        self.assertEqual(headers, ["LOT Nr.", "Breite", "Datum"])

    def test_creates_output_dir(self):
        sub = self.tmp_dir / "sub" / "dir"
        path = create_measurement_file(
            self.process, "REF123", sub, "LOT001", "FA123", "1", date(2026, 4, 1)
        )
        self.assertTrue(sub.exists())
        self.assertTrue(path.exists())


class TestFindExistingFile(unittest.TestCase):

    def setUp(self):
        self.tmp_dir = Path(tempfile.mkdtemp())
        self.process = _make_process()

    def tearDown(self):
        shutil.rmtree(self.tmp_dir, ignore_errors=True)

    def test_not_found(self):
        result = find_existing_file("LOT001", "FA123", "REF123", "IPC1_Test", self.tmp_dir)
        self.assertIsNone(result)

    def test_found(self):
        create_measurement_file(
            self.process, "REF123", self.tmp_dir, "LOT001", "FA123", "1", date(2026, 4, 1)
        )
        result = find_existing_file("LOT001", "FA123", "REF123", "IPC1_Test", self.tmp_dir)
        self.assertIsNotNone(result)
        self.assertTrue(result.exists())

    def test_finds_regardless_of_shift_and_date(self):
        # Erstellt mit Schicht 1, sucht ohne Schicht/Datum
        create_measurement_file(
            self.process, "REF123", self.tmp_dir, "LOT001", "FA123", "2", date(2026, 1, 1)
        )
        result = find_existing_file("LOT001", "FA123", "REF123", "IPC1_Test", self.tmp_dir)
        self.assertIsNotNone(result)

    def test_different_lot_not_found(self):
        create_measurement_file(
            self.process, "REF123", self.tmp_dir, "LOT001", "FA123", "1", date(2026, 4, 1)
        )
        result = find_existing_file("LOT999", "FA123", "REF123", "IPC1_Test", self.tmp_dir)
        self.assertIsNone(result)


class TestCountDataRows(unittest.TestCase):

    def setUp(self):
        self.tmp_dir = Path(tempfile.mkdtemp())
        self.process = _make_process()

    def tearDown(self):
        shutil.rmtree(self.tmp_dir, ignore_errors=True)

    def test_empty_file(self):
        path = create_measurement_file(
            self.process, "REF123", self.tmp_dir, "LOT001", "FA123", "1", date(2026, 4, 1)
        )
        self.assertEqual(count_data_rows(path), 0)

    def test_with_data(self):
        path = create_measurement_file(
            self.process, "REF123", self.tmp_dir, "LOT001", "FA123", "1", date(2026, 4, 1)
        )
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        ws.cell(HEADER_ROW + 1, 1, "LOT-001")
        ws.cell(HEADER_ROW + 2, 1, "LOT-002")
        wb.save(path)
        wb.close()
        self.assertEqual(count_data_rows(path), 2)

    def test_nonexistent_file(self):
        self.assertEqual(count_data_rows(Path("/nonexistent.xlsx")), 0)


class TestWriteInfoHeader(unittest.TestCase):

    def setUp(self):
        self.tmp_dir = Path(tempfile.mkdtemp())
        self.process = _make_process()

    def tearDown(self):
        shutil.rmtree(self.tmp_dir, ignore_errors=True)

    def test_writes_info(self):
        path = create_measurement_file(
            self.process, "REF123", self.tmp_dir, "LOT001", "FA123", "1", date(2026, 4, 1)
        )
        write_info_header(
            path, "Test Product", "IPC1 Test", "1", date(2026, 4, 1),
            extra_info=[("FA-Nr.:", "FA-12345"), ("LOT Nr.:", "LOT001")],
        )
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        # Zeile 1: nur Produktname (fett)
        self.assertEqual(ws.cell(1, 1).value, "Test Product")
        # Linke Spalten ab Zeile 2: fester Kern
        self.assertEqual(ws.cell(2, 1).value, "Prozess:")
        self.assertEqual(ws.cell(2, 2).value, "IPC1 Test")
        self.assertEqual(ws.cell(3, 1).value, "Schicht:")
        self.assertEqual(ws.cell(3, 2).value, "1")
        self.assertEqual(ws.cell(4, 1).value, "Datum:")
        self.assertEqual(ws.cell(4, 2).value, "2026-04-01")
        # Rechte Spalten ab Zeile 2: extra_info
        self.assertEqual(ws.cell(2, 3).value, "FA-Nr.:")
        self.assertEqual(ws.cell(2, 4).value, "FA-12345")
        self.assertEqual(ws.cell(3, 3).value, "LOT Nr.:")
        self.assertEqual(ws.cell(3, 4).value, "LOT001")
        # Column headers in HEADER_ROW (=9)
        self.assertEqual(ws.cell(HEADER_ROW, 1).value, "LOT Nr.")
        wb.close()


if __name__ == "__main__":
    unittest.main()
