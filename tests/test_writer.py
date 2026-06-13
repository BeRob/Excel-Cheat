"""Tests für den Excel-Writer (neue config-basierte API)."""

import os
import shutil
import tempfile
import unittest
from datetime import date
from pathlib import Path

import openpyxl

from src.audit.file_lock import acquire_lock, release_lock
from src.config.process_config import FieldDef, ProcessConfig
from src.config.settings import HEADER_ROW
from src.excel.creator import create_measurement_file
from src.excel.writer import _excel_lock, write_measurement_row, write_measurement_rows


def _make_process() -> ProcessConfig:
    return ProcessConfig(
        template_id="IPC1_Test",
        display_name="IPC1 Test",
        fields=[
            FieldDef(id="lot", display_name="LOT Nr.", type="text", role="context", persistent=True),
            FieldDef(id="rollen", display_name="Rollen Nr.", type="text", role="context"),
            FieldDef(id="breite", display_name="Breite 1", type="number", role="measurement",
                     spec_min=180, spec_max=190),
            FieldDef(id="breite2", display_name="Breite 2", type="number", role="measurement",
                     spec_min=180, spec_max=190),
            FieldDef(id="datum", display_name="Datum", type="text", role="auto"),
            FieldDef(id="bearbeiter", display_name="Bearbeiter", type="text", role="auto"),
        ],
    )


class TestWriteMeasurementRow(unittest.TestCase):

    def setUp(self):
        self.tmp_dir = Path(tempfile.mkdtemp())
        self.process = _make_process()
        self.filepath = create_measurement_file(
            self.process, "TEST", self.tmp_dir, "LOT001", "FA123", "1", date(2026, 4, 2)
        )

    def tearDown(self):
        shutil.rmtree(self.tmp_dir, ignore_errors=True)

    def test_write_appends_row(self):
        result = write_measurement_row(
            self.filepath,
            self.process,
            context_values={"LOT Nr.": "LOT-001", "Rollen Nr.": "R1"},
            measurements={"Breite 1": 185.0, "Breite 2": 186.0},
            auto_values={"Datum": "2026-04-02", "Bearbeiter": "testuser"},
        )
        self.assertTrue(result.success)
        self.assertEqual(result.row_number, HEADER_ROW + 1)

    def test_values_in_correct_columns(self):
        write_measurement_row(
            self.filepath,
            self.process,
            context_values={"LOT Nr.": "LOT-001", "Rollen Nr.": "R1"},
            measurements={"Breite 1": 185.0, "Breite 2": 186.0},
            auto_values={"Datum": "2026-04-02", "Bearbeiter": "testuser"},
        )
        wb = openpyxl.load_workbook(self.filepath, read_only=True)
        ws = wb.active
        data_row = HEADER_ROW + 1
        row = [ws.cell(data_row, c).value for c in range(1, 7)]
        self.assertEqual(row[0], "LOT-001")     # col 1: LOT Nr.
        self.assertEqual(row[1], "R1")           # col 2: Rollen Nr.
        self.assertAlmostEqual(row[2], 185.0)    # col 3: Breite 1
        self.assertAlmostEqual(row[3], 186.0)    # col 4: Breite 2
        self.assertEqual(row[4], "2026-04-02")   # col 5: Datum
        self.assertEqual(row[5], "testuser")     # col 6: Bearbeiter
        wb.close()

    def test_two_sequential_writes(self):
        ctx = {"LOT Nr.": "LOT-001", "Rollen Nr.": "R1"}
        meas = {"Breite 1": 185.0, "Breite 2": 186.0}
        auto = {"Datum": "2026-04-02", "Bearbeiter": "user"}

        r1 = write_measurement_row(self.filepath, self.process, ctx, meas, auto)
        r2 = write_measurement_row(self.filepath, self.process, ctx, meas, auto)
        self.assertEqual(r1.row_number, HEADER_ROW + 1)
        self.assertEqual(r2.row_number, HEADER_ROW + 2)

    def test_file_not_found(self):
        result = write_measurement_row(
            Path("/nonexistent/file.xlsx"),
            self.process,
            {}, {}, {},
        )
        self.assertFalse(result.success)
        self.assertIsNotNone(result.error)

    def test_none_values_skipped(self):
        result = write_measurement_row(
            self.filepath,
            self.process,
            context_values={"LOT Nr.": "LOT-001"},
            measurements={"Breite 1": None},
            auto_values={"Datum": "2026-04-02"},
        )
        self.assertTrue(result.success)
        wb = openpyxl.load_workbook(self.filepath, read_only=True)
        ws = wb.active
        data_row = HEADER_ROW + 1
        # Breite 1 (col 3) should be None since we passed None
        self.assertIsNone(ws.cell(data_row, 3).value)
        # Rollen Nr. (col 2) should be None since not provided
        self.assertIsNone(ws.cell(data_row, 2).value)
        wb.close()

    def test_partial_measurements(self):
        result = write_measurement_row(
            self.filepath,
            self.process,
            context_values={"LOT Nr.": "LOT-001"},
            measurements={"Breite 1": 185.0},
            auto_values={},
        )
        self.assertTrue(result.success)
        wb = openpyxl.load_workbook(self.filepath, read_only=True)
        ws = wb.active
        data_row = HEADER_ROW + 1
        self.assertAlmostEqual(ws.cell(data_row, 3).value, 185.0)
        self.assertIsNone(ws.cell(data_row, 4).value)  # Breite 2 not provided
        wb.close()


class TestWriterGmpHardening(unittest.TestCase):
    """Header-Validierung, Spaltentreffer-Pflicht, Lock und atomares Speichern."""

    def setUp(self):
        self.tmp_dir = Path(tempfile.mkdtemp())
        self.process = _make_process()
        self.filepath = create_measurement_file(
            self.process, "TEST", self.tmp_dir, "LOT001", "FA123", "1", date(2026, 4, 2)
        )

    def tearDown(self):
        shutil.rmtree(self.tmp_dir, ignore_errors=True)

    def _swap_headers(self, name_a: str, name_b: str) -> None:
        wb = openpyxl.load_workbook(self.filepath)
        ws = wb.active
        cols = {ws.cell(HEADER_ROW, c).value: c for c in range(1, ws.max_column + 1)}
        ws.cell(HEADER_ROW, cols[name_a], name_b)
        ws.cell(HEADER_ROW, cols[name_b], name_a)
        wb.save(self.filepath)
        wb.close()

    def test_resume_with_reordered_headers_writes_per_header(self):
        # Resume-Garantie: Spaltenzuordnung folgt der Header-Zeile der Datei,
        # nicht der Feld-Reihenfolge der Config.
        self._swap_headers("Rollen Nr.", "Breite 1")
        result = write_measurement_row(
            self.filepath, self.process,
            context_values={"LOT Nr.": "L1", "Rollen Nr.": "R1"},
            measurements={"Breite 1": 185.0, "Breite 2": 186.0},
            auto_values={},
        )
        self.assertTrue(result.success)
        wb = openpyxl.load_workbook(self.filepath, read_only=True)
        ws = wb.active
        cols = {ws.cell(HEADER_ROW, c).value: c for c in range(1, ws.max_column + 1)}
        self.assertAlmostEqual(
            ws.cell(HEADER_ROW + 1, cols["Breite 1"]).value, 185.0
        )
        self.assertEqual(ws.cell(HEADER_ROW + 1, cols["Rollen Nr."]).value, "R1")
        wb.close()

    def test_missing_header_blocks_write(self):
        wb = openpyxl.load_workbook(self.filepath)
        ws = wb.active
        cols = {ws.cell(HEADER_ROW, c).value: c for c in range(1, ws.max_column + 1)}
        # cell(row, col, None) ist in openpyxl ein No-Op — Wert direkt löschen
        ws.cell(HEADER_ROW, cols["Breite 2"]).value = None
        wb.save(self.filepath)
        wb.close()

        result = write_measurement_row(
            self.filepath, self.process,
            context_values={"LOT Nr.": "L1"},
            measurements={"Breite 1": 185.0},
            auto_values={},
        )
        self.assertFalse(result.success)
        self.assertIn("Breite 2", result.error)

        # Keine Datenzeile geschrieben
        wb = openpyxl.load_workbook(self.filepath, read_only=True)
        self.assertEqual(wb.active.max_row, HEADER_ROW)
        wb.close()

    def test_value_without_column_blocks_write(self):
        # Stiller Feldverlust verboten: ein Wert ohne Spaltentreffer bricht ab.
        result = write_measurement_row(
            self.filepath, self.process,
            context_values={"LOT Nr.": "L1"},
            measurements={"Breite 1": 185.0, "Unbekanntes Feld": 1.0},
            auto_values={},
        )
        self.assertFalse(result.success)
        self.assertIn("Unbekanntes Feld", result.error)

    def test_info_header_values_are_ignored_not_errors(self):
        # info_header-Felder haben bewusst keine Spalte (stehen im Info-Block) —
        # ihre Werte im Kontext-Dict dürfen das Schreiben nicht abbrechen.
        process = ProcessConfig(
            template_id="IPC9_Test",
            display_name="IPC9 Test",
            fields=[
                FieldDef(id="fa_nr", display_name="FA-Nr.", type="text",
                         role="context", persistent=True, info_header=True),
                FieldDef(id="breite", display_name="Breite 1", type="number",
                         role="measurement"),
            ],
        )
        path = create_measurement_file(
            process, "TEST2", self.tmp_dir, "L2", "FA2", "1", date(2026, 4, 2)
        )
        result = write_measurement_row(
            path, process,
            context_values={"FA-Nr.": "FA2"},
            measurements={"Breite 1": 1.0},
            auto_values={},
        )
        self.assertTrue(result.success)

    def test_multi_rows_with_process_validation(self):
        rows = [
            {"LOT Nr.": "L1", "Breite 1": 181.0, "Breite 2": 182.0},
            {"LOT Nr.": "L1", "Breite 1": 183.0, "Breite 2": 184.0},
        ]
        result = write_measurement_rows(self.filepath, rows, process=self.process)
        self.assertTrue(result.success)
        self.assertEqual(result.row_number, HEADER_ROW + 2)

    def test_excel_lock_times_out_when_held(self):
        lock_path = str(self.filepath) + ".lock"
        fd = os.open(lock_path, os.O_RDWR | os.O_CREAT, 0o644)
        self.assertTrue(acquire_lock(fd, 1.0))
        try:
            with _excel_lock(self.filepath, timeout=0.1) as got_lock:
                self.assertFalse(got_lock)
        finally:
            release_lock(fd)
            os.close(fd)

    def test_atomic_save_leaves_no_tmp_file(self):
        result = write_measurement_row(
            self.filepath, self.process,
            context_values={"LOT Nr.": "L1"},
            measurements={"Breite 1": 185.0},
            auto_values={},
        )
        self.assertTrue(result.success)
        self.assertEqual(list(self.tmp_dir.glob("*.tmp~")), [])


if __name__ == "__main__":
    unittest.main()
